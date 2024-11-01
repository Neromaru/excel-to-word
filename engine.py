import os
import pathlib as pl
import uuid
import re
from collections import defaultdict

from docxtpl import DocxTemplate
import openpyxl
from openpyxl.styles.numbers import (
    FORMAT_PERCENTAGE_00,
    FORMAT_DATE_YYYYMMDD2,
    FORMAT_DATE_DDMMYY,
)


class TemplateGenerator(object):
    def __init__(
        self,
        path_to_data_folder=None,
        templates=None,
        where_to_save=None,
        named_header=None,
    ):
        self.data_file = path_to_data_folder
        self.templates = templates
        self.excel = None
        self.save_folder = where_to_save
        self.named_header = named_header
        self._headers = None
        self._group_by_headers: bool = False

    def group_headers_and_values(self, headers, row):
        pattern = r"\b(\w+)(\d+)\b"
        grouped_data = defaultdict(dict)

        for header in headers:
            match = re.search(pattern, header)
            if match:
                base, number = match.groups()
                if row[header].internal_value is not None:
                    grouped_data[number][base] = self.format_cell_value(row[header])

        return [{"number": key, **value} for key, value in grouped_data.items()]

    @property
    def headers(self):
        if self._headers is None:
            self._headers = [i.value for i in list(self.excel.iter_rows())[0]]
        return self._headers

    def set_group_by_headers(self, toggle: bool) -> None:
        self._group_by_headers = toggle

    @staticmethod
    def format_cell_value(cell):
        """Convert cell value to a string formatted according to the cell's number format."""
        format_code = cell.number_format
        value = cell.value

        if not value:
            return value

        if format_code == FORMAT_PERCENTAGE_00:
            # Handle percentage
            return f"{value * 100:.2f}%"
        elif format_code in [FORMAT_DATE_YYYYMMDD2, FORMAT_DATE_DDMMYY] or cell.is_date:
            # Handle date
            return value.strftime("%d.%m.%Y")
        elif isinstance(value, float):
            # Handle general number format as 000 000,00
            return f"{value:,.2f}".replace(",", " ").replace(".", ",")
        else:
            # Handle all other cases
            return str(value)

    def read_data(self):
        extension = pl.Path(self.data_file).suffix
        if extension == ".xlsx":
            self.excel = openpyxl.load_workbook(self.data_file).active
        elif extension == ".xls":
            raise ValueError(
                "Ця версія програми не підритмує старі формати Екселю будь ласка збережіть у форматі .xlsx"
            )
        elif not self.excel:
            raise ValueError("Додано файл не підтримуваного формату")

    def generate_templates(self):
        init_row = 2
        for row in self.excel.iter_rows(min_row=init_row):
            self._make_template(row, init_row)
            init_row += 1

    def _make_template(self, row, row_number):
        row = dict(zip(self.headers, row))
        index_name = pl.Path(f"{str(row_number)}_{str(row[self.named_header].value)}")
        if self._group_by_headers:
            write_folder = pl.Path(self.save_folder) / index_name
            if write_folder.exists():
                write_folder = self.save_folder / pl.Path(
                    index_name.name + str(uuid.uuid4())[:8]
                )
            write_folder.mkdir(parents=True, exist_ok=True)
        else:
            write_folder = pl.Path(self.save_folder)

        for template in self.templates:
            template_docx = DocxTemplate(template)

            fields = {
                variable: self.format_cell_value(row[variable])
                for variable in self.headers
            }
            fields["МНОЖИНИ"] = self.group_headers_and_values(self.headers, row)
            template_docx.render(fields)
            template_basename = f"{index_name.name}_{template.name}"
            template_docx.save(os.path.join(write_folder, template_basename))
